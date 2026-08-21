from __future__ import annotations

import argparse
import hashlib
import re
import sys
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any
from urllib.parse import urlparse, urlunparse

from openpyxl import load_workbook
from psycopg2.extras import Json, execute_batch

PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from DB.db_utils import get_db_cursor


PROVIDER = "CULTURE_FACILITY"
SERVICE_GROUP = "체험"
COLLECTION_CATEGORY = "체험"
SOURCE_NAME = "2025 전국 문화기반시설 총람.xlsx"
DOCUMENT_DIR = PROJECT_ROOT / "document"


@dataclass(frozen=True)
class SheetConfig:
    start_row: int
    sido: int
    sigungu: int
    name: int
    address: int
    phone: int
    homepage: int | None = None
    open_date: int | None = None
    operating_hours: tuple[tuple[str, int], ...] = ()
    regular_holiday: int | None = None
    fees: tuple[tuple[str, int], ...] = ()
    info: tuple[tuple[str, int], ...] = ()


MUSEUM_FEES = (
    ("상설 일반", 49),
    ("상설 청소년", 50),
    ("상설 어린이", 51),
    ("상설 노인", 52),
    ("상설 단체", 53),
    ("상설 할인", 54),
    ("상설 무료대상", 55),
    ("기획 일반", 58),
    ("기획 청소년", 59),
    ("기획 어린이", 60),
    ("기획 노인", 61),
    ("기획 단체", 62),
    ("기획 할인", 63),
    ("기획 무료대상", 64),
)

SHEET_CONFIGS: dict[str, SheetConfig] = {
    "국립도서관": SheetConfig(
        start_row=7,
        sido=2,
        sigungu=3,
        name=5,
        address=6,
        phone=7,
        homepage=8,
        open_date=9,
        info=(("설립연도", 9),),
    ),
    "공공도서관": SheetConfig(
        start_row=7,
        sido=2,
        sigungu=3,
        name=5,
        address=6,
        phone=7,
        homepage=8,
        open_date=9,
        info=(("개관연도", 9),),
    ),
    "박물관": SheetConfig(
        start_row=10,
        sido=2,
        sigungu=3,
        name=6,
        address=7,
        phone=8,
        homepage=13,
        open_date=9,
        operating_hours=(("평일", 44), ("휴일", 45)),
        regular_holiday=46,
        fees=MUSEUM_FEES,
        info=(
            ("운영주체", 4),
            ("구분", 5),
            ("개관일", 9),
            ("교육프로그램", 38),
            ("문화행사", 39),
            ("전시연계프로그램", 40),
            ("체험프로그램", 41),
            ("기타프로그램", 42),
            ("연간개관일", 43),
            ("연간관람객", 47),
            ("일평균관람객", 48),
        ),
    ),
    "미술관": SheetConfig(
        start_row=10,
        sido=2,
        sigungu=3,
        name=6,
        address=7,
        phone=8,
        homepage=13,
        open_date=9,
        operating_hours=(("평일", 44), ("휴일", 45)),
        regular_holiday=46,
        fees=MUSEUM_FEES,
        info=(
            ("운영주체", 4),
            ("구분", 5),
            ("개관일", 9),
            ("교육프로그램", 38),
            ("문화행사", 39),
            ("전시연계프로그램", 40),
            ("체험프로그램", 41),
            ("기타프로그램", 42),
            ("연간개관일", 43),
            ("연간관람객", 47),
            ("일평균관람객", 48),
        ),
    ),
    "생활문화센터": SheetConfig(
        start_row=6,
        sido=2,
        sigungu=3,
        name=4,
        address=9,
        phone=10,
        homepage=11,
        open_date=8,
        operating_hours=(("평일", 12), ("토요일", 13), ("일요일", 14), ("공휴일", 15)),
        regular_holiday=16,
        info=(("운영방식", 5), ("운영주체", 6), ("공간유형", 7), ("개관일", 8), ("연간프로그램", 29), ("연간이용자", 30)),
    ),
    "문예회관": SheetConfig(
        start_row=6,
        sido=2,
        sigungu=3,
        name=5,
        address=6,
        phone=7,
        homepage=10,
        open_date=11,
        info=(("건립주체", 4), ("운영기관", 8), ("운영유형", 9), ("개관일", 11), ("공연장수", 12), ("전시장수", 17), ("문화교육실수", 21)),
    ),
    "지방문화원": SheetConfig(
        start_row=6,
        sido=2,
        sigungu=3,
        name=4,
        address=7,
        phone=8,
        homepage=9,
        open_date=6,
        info=(("원장", 5), ("설립일", 6), ("프로그램수", 27), ("참여자수", 28)),
    ),
    "문화의집": SheetConfig(
        start_row=7,
        sido=2,
        sigungu=3,
        name=4,
        address=5,
        phone=6,
        homepage=7,
        open_date=8,
        operating_hours=(("평일", 17), ("토요일", 18), ("일요일", 19), ("공휴일", 20)),
        regular_holiday=16,
        info=(("개관일", 8), ("운영방식", 9), ("운영주체", 10), ("연간이용자", 12), ("프로그램1", 34), ("프로그램2", 37), ("프로그램3", 40), ("동아리", 43)),
    ),
    "문학관": SheetConfig(
        start_row=9,
        sido=2,
        sigungu=3,
        name=10,
        address=11,
        phone=12,
        homepage=14,
        open_date=13,
        operating_hours=(("평일", 52), ("주말/공휴일", 53)),
        fees=(("성인", 56),),
        info=(("개관일", 13), ("프로그램수", 49), ("연간개관일", 51), ("연간관람객", 54), ("일평균관람객", 55)),
    ),
    "(부록)지역문화재단": SheetConfig(
        start_row=6,
        sido=2,
        sigungu=3,
        name=4,
        address=5,
        phone=6,
        homepage=7,
        open_date=8,
        info=(("설립일", 8), ("주요사업", 17)),
    ),
}


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        value = value.date()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        value = int(value) if value == int(value) else float(value)
    text = str(value).replace("\r", "\n").replace("\t", " ")
    text = re.sub(r"\s*\n+\s*", " / ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return "" if text in {"-", "--", "없음", "해당없음", "미운영", "nan", "None"} else text


def cell(values: tuple[Any, ...], col: int | None) -> str:
    if not col or len(values) < col:
        return ""
    return clean_text(values[col - 1])


def normalize_url(value: str) -> str:
    text = clean_text(value)
    if not text:
        return ""
    for candidate in re.split(r"[,;\s]+", text):
        item = candidate.strip(" ./")
        if not item or "@" in item:
            continue
        if "." not in item and not item.startswith(("http://", "https://")):
            continue
        if item.startswith("www."):
            item = f"https://{item}"
        elif not item.startswith(("http://", "https://")):
            item = f"https://{item}"
        parsed = urlparse(item)
        if parsed.netloc:
            return urlunparse((parsed.scheme, parsed.netloc, parsed.path or "", "", parsed.query or "", ""))
    return ""


def is_summary_row(sido: str, sigungu: str, name: str, address: str) -> bool:
    compact_values = {re.sub(r"\s+", "", value) for value in (sido, sigungu, name)}
    if compact_values & {"계", "합계", "총계", "전국", "소계"}:
        return True
    if not name or name.replace(",", "").replace(".", "").isdigit():
        return True
    if not address and name in {"시설명", "기관명", "명칭"}:
        return True
    return False


def branch_code(sheet: str, name: str, address: str, sido: str, sigungu: str) -> str:
    digest = hashlib.sha1(f"{sheet}|{sido}|{sigungu}|{name}|{address}".encode("utf-8")).hexdigest()[:18].upper()
    return f"CF_{digest}"


def fee_value(value: str) -> str:
    text = clean_text(value)
    if not text:
        return ""
    normalized = text.replace(",", "")
    if re.fullmatch(r"\d+(\.0+)?", normalized):
        amount = int(float(normalized))
        return "무료" if amount == 0 else f"{amount:,}원"
    return text


def summarize_fees(values: tuple[Any, ...], config: SheetConfig) -> str:
    parts: list[str] = []
    for label, col in config.fees:
        value = fee_value(cell(values, col))
        if not value:
            continue
        parts.append(f"{label} {value}")
    if not parts:
        return ""
    if len(parts) > 8:
        return " / ".join(parts[:8]) + f" / 외 {len(parts) - 8}개"
    return " / ".join(parts)


def summarize_hours(values: tuple[Any, ...], config: SheetConfig) -> str:
    parts: list[str] = []
    for label, col in config.operating_hours:
        value = cell(values, col)
        if value:
            parts.append(f"{label} {value}")
    return " / ".join(parts)


def build_basic_info(values: tuple[Any, ...], config: SheetConfig, sheet: str, row: int, sido: str, sigungu: str) -> dict[str, Any]:
    info: dict[str, Any] = {
        "source_file": SOURCE_NAME,
        "source_sheet": sheet,
        "source_row": row,
        "region_sido": sido,
        "region_sigungu": sigungu,
        "classification": COLLECTION_CATEGORY,
    }
    for label, col in config.info:
        value = cell(values, col)
        if value:
            info[label] = value
    fee_details = {label: fee_value(cell(values, col)) for label, col in config.fees}
    fee_details = {label: value for label, value in fee_details.items() if value}
    if fee_details:
        info["요금상세"] = fee_details
    return info


def find_workbook(explicit_path: str | None) -> Path:
    if explicit_path:
        path = Path(explicit_path)
        return path if path.is_absolute() else PROJECT_ROOT / path
    candidates = [
        path
        for path in DOCUMENT_DIR.glob("*.xlsx")
        if all(token in path.name for token in ("문화", "기반", "시설", "총람"))
    ]
    if not candidates:
        raise FileNotFoundError("document 폴더에서 문화기반시설 총람 xlsx 파일을 찾지 못했습니다.")
    return candidates[0]


def parse_workbook(path: Path, limit: int = 0) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows: list[dict[str, Any]] = []
    for sheet in workbook.worksheets:
        config = SHEET_CONFIGS.get(sheet.title)
        if not config:
            continue

        current_sido = ""
        max_col = max(
            [
                config.sido,
                config.sigungu,
                config.name,
                config.address,
                config.phone,
                config.homepage or 0,
                config.open_date or 0,
                config.regular_holiday or 0,
                *(col for _, col in config.operating_hours),
                *(col for _, col in config.fees),
                *(col for _, col in config.info),
            ]
        )
        for row_num, values in enumerate(sheet.iter_rows(min_row=config.start_row, max_col=max_col, values_only=True), start=config.start_row):
            values = tuple(values)
            sido = cell(values, config.sido)
            sigungu = cell(values, config.sigungu)
            name = cell(values, config.name)
            address = cell(values, config.address)
            if sido and re.sub(r"\s+", "", sido) not in {"계", "합계", "총계", "전국", "소계"}:
                current_sido = sido
            elif not sido:
                sido = current_sido
            if is_summary_row(sido, sigungu, name, address):
                continue
            if not address:
                continue

            phone = cell(values, config.phone)
            website_url = normalize_url(cell(values, config.homepage))
            operating_hours = summarize_hours(values, config)
            regular_holiday = cell(values, config.regular_holiday)
            admission_fee = summarize_fees(values, config)
            basic_info = build_basic_info(values, config, sheet.title, row_num, sido, sigungu)

            rows.append(
                {
                    "provider": PROVIDER,
                    "branch_code": branch_code(sheet.title, name, address, sido, sigungu),
                    "name": name[:100],
                    "address": address,
                    "phone": phone or None,
                    "website_url": website_url or None,
                    "operating_hours": operating_hours or None,
                    "facility_type": sheet.title[:80],
                    "facility_category": sheet.title[:80],
                    "facility_source": SOURCE_NAME,
                    "facility_source_sheet": sheet.title,
                    "facility_service_group": SERVICE_GROUP,
                    "facility_collection_category": COLLECTION_CATEGORY,
                    "region_sido": sido or None,
                    "region_sigungu": sigungu or None,
                    "regular_holiday": regular_holiday or None,
                    "admission_fee": admission_fee or None,
                    "basic_info": basic_info,
                    "address_source": f"{SOURCE_NAME}:{sheet.title}:{row_num}",
                }
            )
            if limit and len(rows) >= limit:
                return rows
    return rows


def ensure_schema() -> None:
    with get_db_cursor() as cursor:
        cursor.execute(
            """
            ALTER TABLE branches ADD COLUMN IF NOT EXISTS operating_hours TEXT;
            ALTER TABLE branches ADD COLUMN IF NOT EXISTS website_url TEXT;
            ALTER TABLE branches ADD COLUMN IF NOT EXISTS facility_type VARCHAR(80);
            ALTER TABLE branches ADD COLUMN IF NOT EXISTS facility_category VARCHAR(80);
            ALTER TABLE branches ADD COLUMN IF NOT EXISTS facility_source TEXT;
            ALTER TABLE branches ADD COLUMN IF NOT EXISTS facility_source_sheet TEXT;
            ALTER TABLE branches ADD COLUMN IF NOT EXISTS facility_service_group VARCHAR(50);
            ALTER TABLE branches ADD COLUMN IF NOT EXISTS facility_collection_category VARCHAR(50);
            ALTER TABLE branches ADD COLUMN IF NOT EXISTS region_sido VARCHAR(50);
            ALTER TABLE branches ADD COLUMN IF NOT EXISTS region_sigungu VARCHAR(80);
            ALTER TABLE branches ADD COLUMN IF NOT EXISTS regular_holiday TEXT;
            ALTER TABLE branches ADD COLUMN IF NOT EXISTS admission_fee TEXT;
            ALTER TABLE branches ADD COLUMN IF NOT EXISTS basic_info JSONB DEFAULT '{}'::jsonb;
            ALTER TABLE branches ADD COLUMN IF NOT EXISTS address_source TEXT;
            CREATE INDEX IF NOT EXISTS idx_branches_facility_service_group ON branches(facility_service_group);
            CREATE INDEX IF NOT EXISTS idx_branches_facility_collection_category ON branches(facility_collection_category);
            """
        )


UPSERT_SQL = """
INSERT INTO branches (
    provider, branch_code, name, address, phone, website_url, operating_hours,
    facility_type, facility_category, facility_source, facility_source_sheet,
    facility_service_group, facility_collection_category, region_sido, region_sigungu,
    regular_holiday, admission_fee, basic_info, address_source, location_verified,
    location_checked_at
) VALUES (
    %(provider)s, %(branch_code)s, %(name)s, %(address)s, %(phone)s, %(website_url)s, %(operating_hours)s,
    %(facility_type)s, %(facility_category)s, %(facility_source)s, %(facility_source_sheet)s,
    %(facility_service_group)s, %(facility_collection_category)s, %(region_sido)s, %(region_sigungu)s,
    %(regular_holiday)s, %(admission_fee)s, %(basic_info)s, %(address_source)s, false,
    now()
)
ON CONFLICT (provider, branch_code) DO UPDATE SET
    name = EXCLUDED.name,
    address = EXCLUDED.address,
    phone = COALESCE(EXCLUDED.phone, branches.phone),
    website_url = COALESCE(EXCLUDED.website_url, branches.website_url),
    operating_hours = EXCLUDED.operating_hours,
    facility_type = EXCLUDED.facility_type,
    facility_category = EXCLUDED.facility_category,
    facility_source = EXCLUDED.facility_source,
    facility_source_sheet = EXCLUDED.facility_source_sheet,
    facility_service_group = EXCLUDED.facility_service_group,
    facility_collection_category = EXCLUDED.facility_collection_category,
    region_sido = EXCLUDED.region_sido,
    region_sigungu = EXCLUDED.region_sigungu,
    regular_holiday = EXCLUDED.regular_holiday,
    admission_fee = EXCLUDED.admission_fee,
    basic_info = EXCLUDED.basic_info,
    address_source = EXCLUDED.address_source,
    location_checked_at = now()
"""


def upsert_rows(rows: list[dict[str, Any]], page_size: int) -> int:
    params = [{**row, "basic_info": Json(row["basic_info"])} for row in rows]
    with get_db_cursor() as cursor:
        execute_batch(cursor, UPSERT_SQL, params, page_size=page_size)
    return len(rows)


def print_summary(rows: list[dict[str, Any]]) -> None:
    by_sheet: dict[str, int] = {}
    with_hours = 0
    with_holiday = 0
    with_fee = 0
    for row in rows:
        sheet = row["facility_source_sheet"]
        by_sheet[sheet] = by_sheet.get(sheet, 0) + 1
        with_hours += 1 if row.get("operating_hours") else 0
        with_holiday += 1 if row.get("regular_holiday") else 0
        with_fee += 1 if row.get("admission_fee") else 0

    print(f"parsed={len(rows)} service_group={SERVICE_GROUP} collection_category={COLLECTION_CATEGORY}")
    print(f"with_hours={with_hours} with_holiday={with_holiday} with_fee={with_fee}")
    for sheet, count in sorted(by_sheet.items()):
        print(f"  {sheet}: {count}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import 2025 culture facility directory into branches")
    parser.add_argument("--file", default=None, help="Workbook path. Defaults to document/*문화*기반*시설*총람*.xlsx")
    parser.add_argument("--limit", type=int, default=0, help="Maximum rows to parse/import. 0 means all rows.")
    parser.add_argument("--page-size", type=int, default=500)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--skip-schema", action="store_true")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    workbook_path = find_workbook(args.file)
    rows = parse_workbook(workbook_path, limit=args.limit)
    print(f"source={workbook_path}")
    print_summary(rows)
    if args.dry_run:
        return 0
    if not args.skip_schema:
        ensure_schema()
    imported = upsert_rows(rows, page_size=args.page_size)
    print(f"imported={imported} provider={PROVIDER}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
