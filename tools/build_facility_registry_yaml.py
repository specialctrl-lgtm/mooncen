from __future__ import annotations

import hashlib
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

import yaml
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DOCUMENT_DIR = ROOT / "document"
SPORTS_FILE = DOCUMENT_DIR / "2025 전국 공공체육시설 현황(종합본).xlsx"
CULTURE_FILE = DOCUMENT_DIR / "2025 전국 문화기반시설 총람.xlsx"
OUTPUT_FILE = ROOT / "config" / "facility_registry_crawl_targets.yaml"

SKIP_SPORTS_SHEETS = {
    "표지",
    "일반개요",
    "공공체육시설의분류기준",
    "연도별현황",
    "설치주체별현황",
    "시도별현황",
}

CATEGORY_CODES = {
    "국립도서관": "NATIONAL_LIBRARY",
    "공공도서관": "PUBLIC_LIBRARY",
    "박물관": "MUSEUM",
    "미술관": "ART_MUSEUM",
    "생활문화센터": "LIVING_CULTURE_CENTER",
    "문예회관": "ARTS_CENTER",
    "지방문화원": "LOCAL_CULTURE_CENTER",
    "문화의집": "CULTURE_HOUSE",
    "문학관": "LITERATURE_MUSEUM",
    "(부록)지역문화재단": "CULTURE_FOUNDATION",
}

CULTURE_SHEET_COLUMNS = {
    "국립도서관": {"start_row": 6, "sido": 2, "sigungu": 3, "name": 5, "address": 6, "phone": 7, "homepage": 8},
    "공공도서관": {"start_row": 6, "sido": 2, "sigungu": 3, "name": 5, "address": 6, "phone": 7, "homepage": 8},
    "박물관": {"start_row": 10, "sido": 2, "sigungu": 3, "name": 6, "address": 7, "phone": 8, "homepage": 13},
    "미술관": {"start_row": 10, "sido": 2, "sigungu": 3, "name": 6, "address": 7, "phone": 8, "homepage": 13},
    "생활문화센터": {"start_row": 6, "sido": 2, "sigungu": 3, "name": 4, "address": 9, "phone": 10, "homepage": 11},
    "문예회관": {"start_row": 6, "sido": 2, "sigungu": 3, "name": 5, "address": 6, "phone": 7, "homepage": 10},
    "지방문화원": {"start_row": 6, "sido": 2, "sigungu": 3, "name": 4, "address": 7, "phone": 8, "homepage": 9},
    "문화의집": {"start_row": 7, "sido": 2, "sigungu": 3, "name": 4, "address": 5, "phone": 6, "homepage": 7},
    "문학관": {"start_row": 9, "sido": 2, "sigungu": 3, "name": 10, "address": 11, "phone": 12, "homepage": 14},
    "(부록)지역문화재단": {"start_row": 6, "sido": 2, "sigungu": 3, "name": 4, "address": 5, "phone": 6, "homepage": 7},
}

LOW_VALUE_URL_TOKENS = {
    "facebook.com",
    "instagram.com",
    "twitter.com",
    "youtube.com",
    "blog.naver.com",
    "tv.naver.com",
}
EXCLUDED_URL_TOKENS = {
    "e-ncom.co.kr",
}


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r", "\n").replace("\t", " ")
    text = re.sub(r"\s*\n+\s*", " / ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text.strip(" -–—")


def clean_category(sheet_name: str) -> str:
    return re.sub(r"^\d+\.", "", sheet_name).strip()


def code_token(value: str, default: str) -> str:
    token = re.sub(r"[^A-Z0-9]+", "_", value.upper()).strip("_")
    return token or default


def stable_provider(prefix: str, category_code: str, name: str, address: str, sheet: str, row: int) -> str:
    digest = hashlib.sha1(f"{sheet}|{row}|{name}|{address}".encode("utf-8")).hexdigest()[:10].upper()
    provider = f"{prefix}_{category_code}_{digest}"
    return provider[:50]


def first_homepage(raw: str) -> tuple[str, list[str]]:
    if not raw:
        return "", []
    text = raw.replace("\n", " ")
    candidates = re.split(r"[,;\s]+", text)
    urls: list[str] = []
    for candidate in candidates:
        item = candidate.strip(" ./")
        if not item or item in {"-", "없음", "무", "x", "X"}:
            continue
        if "@" in item and "." in item:
            continue
        if "." not in item and not item.startswith(("http://", "https://")):
            continue
        if item == "http:" or item == "https:":
            continue
        if item.startswith("www."):
            item = "https://" + item
        elif not item.startswith(("http://", "https://")):
            item = "https://" + item
        if any(token in item.lower() for token in EXCLUDED_URL_TOKENS):
            continue
        if item not in urls:
            urls.append(item)

    primary = ""
    for url in urls:
        lowered = url.lower()
        if not any(token in lowered for token in LOW_VALUE_URL_TOKENS):
            primary = url
            break
    if not primary and urls:
        primary = urls[0]
    return primary, urls


def is_summary_row(sido: str, sigungu: str, name: str) -> bool:
    summary_tokens = {"계", "소계", "소 계", "소  계", "소   계", "합계", "전국"}
    if sigungu.replace(" ", "") in {"계", "소계"}:
        return True
    if sido.replace(" ", "") in {"전국", "계", "소계"}:
        return True
    if name.replace(",", "").replace(".", "").isdigit():
        return True
    if name in summary_tokens:
        return True
    return False


def find_sports_columns(ws) -> Optional[dict[str, int]]:
    columns: dict[str, int] = {}
    labels = {
        "sido": {"시도"},
        "sigungu": {"시군구", "시ㆍ군ㆍ구", "시·군·구", "시/군/구"},
        "address": {"주소"},
        "name": {"시설명"},
        "homepage": {"홈페이지주소"},
        "phone": {"운영조직/(연락처)", "운영조직\n(연락처)"},
        "owner": {"관리주체"},
    }
    for values in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 5), max_col=min(ws.max_column, 120), values_only=True):
        for index, value in enumerate(values, start=1):
            text = clean_text(value)
            for key, names in labels.items():
                if key not in columns and text in names:
                    columns[key] = index
    if not {"sido", "sigungu", "address", "name"}.issubset(columns):
        return None
    return columns


def make_target(
    *,
    provider: str,
    name: str,
    branch: str,
    category: str,
    source: str,
    url: str,
    homepage_urls: list[str],
    address: str,
    sido: str,
    sigungu: str,
    phone: str,
    owner: str,
    source_file: str,
    sheet: str,
    row: int,
    priority: int,
    keywords: list[str],
) -> dict[str, Any]:
    region = " ".join(part for part in [sido, sigungu] if part)
    search_base = " ".join(part for part in [region, name] if part)
    target = {
        "provider": provider,
        "name": name,
        "branch": branch or name,
        "category": category,
        "source": source,
        "priority": priority,
        "region": region,
        "address": address,
        "url": url,
        "homepage_urls": homepage_urls,
        "needs_url_discovery": not bool(url),
        "search_queries": [f"{search_base} {keyword}".strip() for keyword in keywords if search_base],
        "phone": phone,
        "owner": owner,
        "registry_source": {
            "file": source_file,
            "sheet": sheet,
            "row": row,
        },
    }
    if source == "cultural_facilities_directory_2025":
        target.update(
            {
                "service_group": "체험",
                "collection_category": "체험",
                "domain_category": "체험",
                "program_type": "체험",
            }
        )
    return {key: value for key, value in target.items() if value not in ("", [], None)}


def load_sports_targets() -> list[dict[str, Any]]:
    wb = load_workbook(SPORTS_FILE, read_only=True, data_only=True)
    targets: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        if ws.title in SKIP_SPORTS_SHEETS:
            continue
        columns = find_sports_columns(ws)
        if not columns:
            continue
        category = clean_category(ws.title)
        category_code = code_token(re.sub(r"^\d+", "", ws.title), "SPORTS")
        current_sido = ""
        max_needed_col = max(columns.values())
        for row, values in enumerate(ws.iter_rows(min_row=6, max_col=max_needed_col, values_only=True), start=6):
            values = tuple(values)
            sido = clean_text(values[columns["sido"] - 1] if len(values) >= columns["sido"] else "")
            sigungu = clean_text(values[columns["sigungu"] - 1] if len(values) >= columns["sigungu"] else "")
            address = clean_text(values[columns["address"] - 1] if len(values) >= columns["address"] else "")
            name = clean_text(values[columns["name"] - 1] if len(values) >= columns["name"] else "")
            if sido and sido.replace(" ", "") not in {"전국", "계", "소계"}:
                current_sido = sido
            if not sido:
                sido = current_sido
            if not name or is_summary_row(sido, sigungu, name):
                continue
            homepage_col = columns.get("homepage", 0)
            phone_col = columns.get("phone", 0)
            owner_col = columns.get("owner", 0)
            homepage_raw = clean_text(values[homepage_col - 1] if homepage_col and len(values) >= homepage_col else "")
            url, homepage_urls = first_homepage(homepage_raw)
            phone = clean_text(values[phone_col - 1] if phone_col and len(values) >= phone_col else "")
            owner = clean_text(values[owner_col - 1] if owner_col and len(values) >= owner_col else "")
            provider = stable_provider("SPORTS", category_code, name, address, ws.title, row)
            targets.append(
                make_target(
                    provider=provider,
                    name=name,
                    branch=name,
                    category=f"공공체육시설/{category}",
                    source="public_sports_facilities_2025",
                    url=url,
                    homepage_urls=homepage_urls,
                    address=address,
                    sido=sido,
                    sigungu=sigungu,
                    phone=phone,
                    owner=owner,
                    source_file=SPORTS_FILE.name,
                    sheet=ws.title,
                    row=row,
                    priority=4 if url else 6,
                    keywords=["수강신청", "강좌신청", "체육 프로그램", "예약"],
                )
            )
    return targets


def load_culture_targets() -> list[dict[str, Any]]:
    wb = load_workbook(CULTURE_FILE, read_only=True, data_only=True)
    targets: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        config = CULTURE_SHEET_COLUMNS.get(ws.title)
        if not config:
            continue
        category_code = CATEGORY_CODES[ws.title]
        max_needed_col = max(config.values())
        for row, values in enumerate(ws.iter_rows(min_row=config["start_row"], max_col=max_needed_col, values_only=True), start=config["start_row"]):
            values = tuple(values)
            sido = clean_text(values[config["sido"] - 1] if len(values) >= config["sido"] else "")
            sigungu = clean_text(values[config["sigungu"] - 1] if len(values) >= config["sigungu"] else "")
            name = clean_text(values[config["name"] - 1] if len(values) >= config["name"] else "")
            address = clean_text(values[config["address"] - 1] if len(values) >= config["address"] else "")
            if not name or is_summary_row(sido, sigungu, name):
                continue
            homepage_raw = clean_text(values[config["homepage"] - 1] if len(values) >= config["homepage"] else "")
            url, homepage_urls = first_homepage(homepage_raw)
            phone = clean_text(values[config["phone"] - 1] if len(values) >= config["phone"] else "")
            provider = stable_provider("CULTURE", category_code, name, address, ws.title, row)
            targets.append(
                make_target(
                    provider=provider,
                    name=name,
                    branch=name,
                    category=f"문화기반시설/{ws.title}",
                    source="cultural_facilities_directory_2025",
                    url=url,
                    homepage_urls=homepage_urls,
                    address=address,
                    sido=sido,
                    sigungu=sigungu,
                    phone=phone,
                    owner="",
                    source_file=CULTURE_FILE.name,
                    sheet=ws.title,
                    row=row,
                    priority=3 if url else 5,
                    keywords=["교육", "체험", "강좌", "문화행사", "전시", "공연", "관람", "예매", "예약"],
                )
            )
    return targets


def summarize(targets: list[dict[str, Any]]) -> dict[str, Any]:
    by_source: dict[str, int] = {}
    by_category: dict[str, int] = {}
    runnable = 0
    for target in targets:
        by_source[target["source"]] = by_source.get(target["source"], 0) + 1
        by_category[target["category"]] = by_category.get(target["category"], 0) + 1
        if target.get("url"):
            runnable += 1
    return {
        "targets": len(targets),
        "runnable_with_url": runnable,
        "needs_url_discovery": len(targets) - runnable,
        "by_source": dict(sorted(by_source.items())),
        "by_category": dict(sorted(by_category.items())),
    }


def main() -> int:
    sports_targets = load_sports_targets()
    culture_targets = load_culture_targets()
    targets = sorted(
        [*sports_targets, *culture_targets],
        key=lambda item: (item["priority"], item["source"], item["category"], item["region"], item["name"]),
    )
    data = {
        "version": 1,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source_files": [SPORTS_FILE.as_posix(), CULTURE_FILE.as_posix()],
        "purpose": "2025 전국 공공체육시설 현황과 2025 전국 문화기반시설 총람 기반 강좌/체험/예약 크롤링 후보 레지스트리",
        "summary": summarize(targets),
        "targets": targets,
    }
    OUTPUT_FILE.write_text(yaml.safe_dump(data, allow_unicode=True, sort_keys=False, width=160), encoding="utf-8")
    print(f"output={OUTPUT_FILE}")
    print(f"targets={data['summary']['targets']}")
    print(f"runnable_with_url={data['summary']['runnable_with_url']}")
    print(f"needs_url_discovery={data['summary']['needs_url_discovery']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
